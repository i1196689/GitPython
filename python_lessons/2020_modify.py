from openpyxl import load_workbook
wb=load_workbook("刘俊空-时间统计表.xlsx")
ws=wb["刘俊空"]
data=ws["B3:EK14"]

#print(ws["A3"].value)
for i in data:
    for j in i:
        if j.value!=1:#单元格的值不为1 则填0
            j.value=0
wb.save("刘俊空-时间统计表.xlsx")